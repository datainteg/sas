"""
Professional SAS Code Analyzer
=============================
A comprehensive tool for analyzing SAS code files and generating detailed Excel reports
with professional styling and formatting.

Features:
- Comprehensive SAS code analysis
- Professional Excel report generation
- Snowflake table detection
- SQL query analysis
- Function block identification
- Macro analysis and tracking
- Dataset creation type detection (Temporary/Permanent)
- Detailed PROC step analysis with parameters
- Enhanced DATA step operations tracking

Author: Akshay Thakare
Version: 1.2
Date: 2025-08-23
"""

import re
import os
from collections import defaultdict
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =====================================================
# PROFESSIONAL STYLING CONFIGURATION
# =====================================================

PROFESSIONAL_COLORS = {
    'primary_blue': '1F4E79',
    'secondary_blue': '2F75B5', 
    'light_blue': 'D9E2F3',
    'accent_green': '70AD47',
    'accent_orange': 'C65911',
    'light_gray': 'F2F2F2',
    'white': 'FFFFFF',
    'dark_text': '1C1C1C',
    'light_text': 'FFFFFF'
}

EXCEL_STYLES = {
    'main_header': {
        'font': Font(name='Calibri', size=16, bold=True, color=PROFESSIONAL_COLORS['light_text']),
        'fill': PatternFill(start_color=PROFESSIONAL_COLORS['primary_blue'], 
                           end_color=PROFESSIONAL_COLORS['primary_blue'], fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center')
    },
    'section_header': {
        'font': Font(name='Calibri', size=14, bold=True, color=PROFESSIONAL_COLORS['light_text']),
        'fill': PatternFill(start_color=PROFESSIONAL_COLORS['secondary_blue'], 
                           end_color=PROFESSIONAL_COLORS['secondary_blue'], fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center')
    },
    'table_header': {
        'font': Font(name='Calibri', size=11, bold=True, color=PROFESSIONAL_COLORS['dark_text']),
        'fill': PatternFill(start_color=PROFESSIONAL_COLORS['light_blue'], 
                           end_color=PROFESSIONAL_COLORS['light_blue'], fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center')
    },
    'data_cell': {
        'font': Font(name='Calibri', size=10, color=PROFESSIONAL_COLORS['dark_text']),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    },
    'data_cell_alt': {
        'font': Font(name='Calibri', size=10, color=PROFESSIONAL_COLORS['dark_text']),
        'fill': PatternFill(start_color=PROFESSIONAL_COLORS['light_gray'], 
                           end_color=PROFESSIONAL_COLORS['light_gray'], fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    }
}

# Border styling for professional appearance
BORDER_THIN = Border(
    left=Side(style='thin', color='808080'),
    right=Side(style='thin', color='808080'),
    top=Side(style='thin', color='808080'),
    bottom=Side(style='thin', color='808080')
)

BORDER_THICK = Border(
    left=Side(style='medium', color=PROFESSIONAL_COLORS['primary_blue']),
    right=Side(style='medium', color=PROFESSIONAL_COLORS['primary_blue']),
    top=Side(style='medium', color=PROFESSIONAL_COLORS['primary_blue']),
    bottom=Side(style='medium', color=PROFESSIONAL_COLORS['primary_blue'])
)


# =====================================================
# CORE ANALYSIS FUNCTIONS
# =====================================================

def initialize_analysis():
    """
    Initialize and return the analysis results data structure and state tracking.
    
    Returns:
        tuple: (results dict, state dict) for tracking analysis progress
    """
    results = {
        'function_blocks': [],
        'datasets_created': defaultdict(list),
        'datasets_used': defaultdict(list),
        'procedures_used': defaultdict(list),
        'macros_defined': defaultdict(dict),
        'macros_called': defaultdict(list),
        'libraries_defined': defaultdict(list),
        'sql_statements': defaultdict(list),
        'variables_used': defaultdict(list),
        'file_operations': defaultdict(list),
        'control_structures': defaultdict(list),
        'include_files': defaultdict(list),
        'system_functions': defaultdict(list),
        'call_routines': defaultdict(list),
        'formats': defaultdict(list),
        'hash_objects': defaultdict(list),
        'ods_statements': defaultdict(list),
        'jdbc_connections': defaultdict(list),
        'snowflake_tables': defaultdict(list),
        'snowflake_queries': [],
        'create_table_info': [],  # Store created tables and types
        'proc_import_details': [],  # Store PROC IMPORT details
        'proc_export_details': [],  # Store PROC EXPORT details
        'data_step_details': [],   # Store DATA step details
        'timeframe_start': None,
        'timeframe_end': None,
        'code_complexity': {},
        'line_analysis': {}
    }
    
    state = {
        'current_blocks': [],
        'include_stack': [],
        'in_sql_block': False,
        'current_sql_query': [],
        'sql_block_start': None,
        'macro_stack': []
    }
    
    return results, state


def clean_line(line):
    """
    Remove block and line comments from SAS code and convert to uppercase.
    
    Args:
        line (str): Raw SAS code line
        
    Returns:
        str: Cleaned and uppercased line
    """
    # Remove block comments /* ... */
    line = re.sub(r'/\*.*?\*/', '', line)
    # Remove line comments starting with *
    line = re.sub(r'^\s*\*.*$', '', line)
    return line.upper()


def classify_line(line):
    """
    Classify a SAS code line into its type category.
    
    Args:
        line (str): Cleaned SAS code line
        
    Returns:
        str: Line classification type
    """
    line = line.strip()
    
    if re.match(r'^\s*%INCLUDE\s+', line):
        return 'INCLUDE'
    elif re.match(r'^\s*DATA\s+', line):
        return 'DATA_STEP'
    elif re.match(r'^\s*PROC\s+', line):
        return 'PROCEDURE'
    elif re.match(r'^\s*%MACRO\s+', line):
        return 'MACRO_DEF'
    elif re.match(r'^\s*%', line):
        return 'MACRO_CALL'
    elif re.match(r'^\s*LIBNAME\s+', line):
        return 'LIBRARY'
    elif re.match(r'^\s*ODS\s+', line):
        return 'ODS'
    elif 'RUN;' in line or 'QUIT;' in line:
        return 'TERMINATOR'
    else:
        return 'STATEMENT'


# =====================================================
# COMPREHENSIVE ANALYSIS FUNCTIONS
# =====================================================

def is_sas_keyword(word):
    """
    Check if a word is a SAS keyword that should be excluded from table detection.
    
    Args:
        word (str): Word to check
        
    Returns:
        bool: True if word is a SAS keyword
    """
    sas_keywords = {
        'SELECT', 'FROM', 'WHERE', 'GROUP', 'BY', 'ORDER', 'HAVING', 'UNION',
        'JOIN', 'INNER', 'LEFT', 'RIGHT', 'FULL', 'OUTER', 'ON', 'AS',
        'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'DROP', 'ALTER', 'TABLE',
        'VIEW', 'INDEX', 'DATABASE', 'SCHEMA', 'COLUMN', 'PRIMARY', 'KEY',
        'FOREIGN', 'CONSTRAINT', 'NOT', 'NULL', 'DEFAULT', 'AUTO_INCREMENT',
        'DISTINCT', 'COUNT', 'SUM', 'AVG', 'MIN', 'MAX', 'AND', 'OR', 'IN',
        'BETWEEN', 'LIKE', 'IS', 'EXISTS', 'CASE', 'WHEN', 'THEN', 'ELSE', 'END'
    }
    return word.upper() in sas_keywords


def extract_tables_from_query(query_text):
    """
    Extract table names from a SQL query string.
    
    Args:
        query_text (str): SQL query text
        
    Returns:
        list: List of table names found in the query
    """
    tables = []
    # Fixed regex patterns with proper escaping
    patterns = [
        r'FROM\s+([A-Z_][A-Z0-9_.]*)',
        r'JOIN\s+([A-Z_][A-Z0-9_.]*)',
        r'INSERT\s+INTO\s+([A-Z_][A-Z0-9_.]*)',
        r'UPDATE\s+([A-Z_][A-Z0-9_.]*)',
        r'CREATE\s+(?:OR\s+REPLACE\s+)?(?:TABLE|VIEW)\s+([A-Z_][A-Z0-9_.]*)'
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, query_text.upper())
        for table in matches:
            if not is_sas_keyword(table) and table not in tables:
                tables.append(table)
    
    return tables


def extract_created_tables_from_sql(sql_text):
    """
    Extract tables created by PROC SQL create statements.

    Args:
        sql_text (str): SQL query text.

    Returns:
        list: List of created tables.
    """
    created_tables = []
    pattern = re.compile(r'CREATE\s+(?:OR\s+REPLACE\s+)?TABLE\s+([A-Z_][A-Z0-9_.]*)', re.IGNORECASE)
    matches = pattern.findall(sql_text)
    for tbl in matches:
        created_tables.append(tbl.upper())
    return created_tables


def determine_dataset_type(table_name):
    """
    Determine if a SAS dataset is temporary or permanent based on its libref.

    Args:
        table_name (str): Dataset name with optional libref.

    Returns:
        str: 'Temporary' or 'Permanent'
    """
    # Split by dot if libref present
    parts = table_name.split('.')
    if len(parts) == 1:
        # No libref means temporary (WORK)
        return 'Temporary'
    libref = parts[0].upper()
    if libref in ['WORK', '']:
        return 'Temporary'
    return 'Permanent'


def extract_proc_options(lines_dict, start_line, end_line, proc_type):
    """
    Extract options from PROC IMPORT/EXPORT blocks.
    
    Args:
        lines_dict (dict): Dictionary of line_num -> line_info
        start_line (int): Block start line
        end_line (int): Block end line
        proc_type (str): 'IMPORT' or 'EXPORT'
        
    Returns:
        dict: Dictionary of extracted options
    """
    options = {'OUT': '', 'DBMS': '', 'DATAFILE': '', 'DATA': '', 'OUTFILE': ''}
    
    for line_num in range(start_line, end_line + 1):
        if line_num in lines_dict:
            line = lines_dict[line_num].get('original', '').upper()
            
            # Extract options using regex
            if 'OUT=' in line:
                match = re.search(r'OUT\s*=\s*([^\s;]+)', line)
                if match:
                    options['OUT'] = match.group(1).strip()
                    
            if 'DATA=' in line:
                match = re.search(r'DATA\s*=\s*([^\s;]+)', line)
                if match:
                    options['DATA'] = match.group(1).strip()
                    
            if 'DBMS=' in line:
                match = re.search(r'DBMS\s*=\s*([^\s;]+)', line)
                if match:
                    options['DBMS'] = match.group(1).strip()
                    
            if 'DATAFILE=' in line:
                match = re.search(r'DATAFILE\s*=\s*["\']?([^"\';\s]+)["\']?', line)
                if match:
                    options['DATAFILE'] = match.group(1).strip()
                    
            if 'OUTFILE=' in line:
                match = re.search(r'OUTFILE\s*=\s*["\']?([^"\';\s]+)["\']?', line)
                if match:
                    options['OUTFILE'] = match.group(1).strip()
    
    return options


def extract_data_step_operations(lines_dict, start_line, end_line):
    """
    Extract operations used in a DATA step.
    
    Args:
        lines_dict (dict): Dictionary of line_num -> line_info
        start_line (int): Block start line
        end_line (int): Block end line
        
    Returns:
        list: List of operations found
    """
    operations = []
    
    for line_num in range(start_line, end_line + 1):
        if line_num in lines_dict:
            line = lines_dict[line_num].get('original', '').upper()
            
            if re.search(r'\bSET\s+', line):
                operations.append('SET')
            if re.search(r'\bMERGE\s+', line):
                operations.append('MERGE')
            if re.search(r'\bBY\s+', line):
                operations.append('BY')
            if re.search(r'\bRETAIN\s+', line):
                operations.append('RETAIN')
            if re.search(r'\bIF\s+.*\bTHEN\b', line):
                operations.append('IF-THEN')
            if re.search(r'\bELSE\b', line):
                operations.append('ELSE')
            if re.search(r'\bDO\s+', line):
                operations.append('DO')
            if re.search(r'\bOUTPUT\s*;', line):
                operations.append('OUTPUT')
            if re.search(r'\bFORMAT\s+', line):
                operations.append('FORMAT')
            if re.search(r'\bLABEL\s+', line):
                operations.append('LABEL')
            if re.search(r'\bLENGTH\s+', line):
                operations.append('LENGTH')
            if re.search(r'\bARRAY\s+', line):
                operations.append('ARRAY')
    
    # Remove duplicates and return
    return list(set(operations))


def analyze_snowflake_references(line, line_num, results, state):
    """
    Detect Snowflake-specific table references and SQL patterns.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
        state (dict): Current analysis state
    """
    line_upper = line.upper()
    
    # Track PROC SQL blocks
    if re.search(r'^\s*PROC\s+SQL', line_upper):
        state['in_sql_block'] = True
        state['sql_block_start'] = line_num
        state['current_sql_query'] = []
        return
    
    # Track end of SQL blocks
    if state['in_sql_block'] and ('QUIT;' in line_upper or 'RUN;' in line_upper):
        if state['current_sql_query']:
            query_text = ' '.join(state['current_sql_query'])
            
            # Extract created tables from query
            created_tables = extract_created_tables_from_sql(query_text)
            created_table_name = created_tables[0] if created_tables else 'None'
            created_table_type = determine_dataset_type(created_table_name) if created_table_name != 'None' else 'N/A'
            
            results['snowflake_queries'].append({
                'start_line': state['sql_block_start'],
                'end_line': line_num,
                'query': query_text,
                'tables_referenced': extract_tables_from_query(query_text),
                'created_table_name': created_table_name,
                'created_table_type': created_table_type
            })
            
            # Store created table info
            for tbl in created_tables:
                tbl_type = determine_dataset_type(tbl)
                results['create_table_info'].append((tbl, tbl_type, state['sql_block_start']))
                
        state['in_sql_block'] = False
        state['current_sql_query'] = []
        state['sql_block_start'] = None
        return
    
    # Collect SQL query lines
    if state['in_sql_block'] and line.strip():
        if not re.match(r'^\s*PROC\s+SQL', line_upper) and not ('QUIT;' in line_upper or 'RUN;' in line_upper):
            state['current_sql_query'].append(line.strip())
    
    # Detect Snowflake table patterns
    snowflake_patterns = [
        r'\b([A-Z_][A-Z0-9_]*\.[A-Z_][A-Z0-9_]*\.[A-Z_][A-Z0-9_]*)\b',
        r'\b([A-Z_][A-Z0-9_]*\.[A-Z_][A-Z0-9_]*)\b',
        r'(?:FROM|JOIN)\s+([A-Z_][A-Z0-9_.]*)',
        r'(?:INSERT\s+INTO|UPDATE)\s+([A-Z_][A-Z0-9_.]*)',
        r'CREATE\s+(?:OR\s+REPLACE\s+)?(?:TABLE|VIEW)\s+([A-Z_][A-Z0-9_.]*)'
    ]
    
    for pattern in snowflake_patterns:
        matches = re.findall(pattern, line_upper)
        for table_name in matches:
            if not is_sas_keyword(table_name):
                results['snowflake_tables'][table_name].append(line_num)


def analyze_include_files(line, line_num, results):
    """
    Analyze and extract %INCLUDE file references from SAS code.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
    """
    include_pattern = re.compile(
        r'%INCLUDE\s*["\']([^"\']+)["\']|%INCLUDE\s*([^;]+);?',
        re.IGNORECASE
    )
    
    matches = include_pattern.findall(line)
    for quoted, unquoted in matches:
        filename = quoted if quoted else unquoted
        filename = filename.strip('"').strip("'").rstrip(';').strip()
        if filename:
            results['include_files'][filename].append(line_num)


def analyze_system_functions(line, line_num, results):
    """
    Detect SAS system functions used in the code.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
    """
    sas_functions = [
        # Numeric functions
        'SUM', 'MEAN', 'MIN', 'MAX', 'COUNT', 'N', 'NMISS',
        'ROUND', 'CEIL', 'FLOOR', 'INT', 'ABS', 'LOG', 'EXP', 'SQRT',
        'SIN', 'COS', 'TAN', 'RAND', 'RANUNI', 'NORMAL', 'GAMMA', 'BETA',
        
        # Character functions
        'SUBSTR', 'TRIM', 'STRIP', 'LEFT', 'RIGHT', 'LENGTH',
        'UPCASE', 'LOWCASE', 'PROPCASE', 'COMPRESS', 'TRANSLATE',
        'INDEX', 'FIND', 'SCAN', 'CATS', 'CATX', 'CAT',
        
        # Date/Time functions
        'TODAY', 'DATE', 'DATETIME', 'TIME', 'DATEPART', 'TIMEPART',
        'YEAR', 'MONTH', 'DAY', 'WEEKDAY', 'MDY', 'YMD',
        'INTCK', 'INTNX', 'DATDIF', 'JULDATE',
        
        # Other functions
        'INPUT', 'PUT', 'COALESCEC', 'COALESCE', 'IFC', 'IFN', 'MISSING'
    ]
    
    for func in sas_functions:
        pattern = rf'\b{func}\s*\('
        if re.search(pattern, line, re.IGNORECASE):
            results['system_functions'][func].append(line_num)


def analyze_procedures(line, line_num, results):
    """
    Analyze PROC statements and their data sources.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
    """
    proc_match = re.search(r'PROC\s+([A-Z]+)(?:\s+DATA\s*=\s*([A-Z_][A-Z0-9_.]*))?\s*;?', line, re.IGNORECASE)
    if proc_match:
        proc_name = proc_match.group(1).upper()
        dataset = proc_match.group(2).upper() if proc_match.group(2) else 'UNKNOWN'
        results['procedures_used'][proc_name].append({
            'line': line_num,
            'dataset': dataset
        })

def analyze_data_operations(line, line_num, results):
    """
    Analyze DATA step operations and dataset usage.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
    """
    # Analyze DATA statement for dataset creation
    data_match = re.search(r'DATA\s+([A-Z_][A-Z0-9_.]*(?:\s+[A-Z_][A-Z0-9_.]*)*)', line, re.IGNORECASE)
    if data_match:
        datasets = data_match.group(1).split()
        for dataset in datasets:
            dataset_up = dataset.upper()
            results['datasets_created'][dataset_up].append(line_num)
            # Store create table + type
            dtype = determine_dataset_type(dataset_up)
            results['create_table_info'].append((dataset_up, dtype, line_num))
    
    # Analyze dataset usage patterns
    usage_patterns = [
        ('SET', r'SET\s+([A-Z_][A-Z0-9_.]*(?:\s+[A-Z_][A-Z0-9_.]*)*?)(?:\s|;|$)'),
        ('MERGE', r'MERGE\s+([A-Z_][A-Z0-9_.]*(?:\s+[A-Z_][A-Z0-9_.]*)*?)(?:\s|;|$)'),
        ('UPDATE', r'UPDATE\s+([A-Z_][A-Z0-9_.]*(?:\s+[A-Z_][A-Z0-9_.]*)*?)(?:\s|;|$)')
    ]
    
    for operation, pattern in usage_patterns:
        matches = re.findall(pattern, line, re.IGNORECASE)
        for match in matches:
            datasets = match.split()
            for dataset in datasets:
                results['datasets_used'][f"{operation}_{dataset.upper()}"].append(line_num)


# def analyze_macros(line, line_num, results):
#     """
#     Analyze macro definitions and calls.
    
#     Args:
#         line (str): Code line to analyze
#         line_num (int): Line number
#         results (dict): Results dictionary to update
#     """
#     # Analyze macro definitions
#     macro_def_match = re.search(r'%MACRO\s+([A-Z_][A-Z0-9_]*)\(([^)]*)\)?\s*;?', line, re.IGNORECASE)
#     if macro_def_match:
#         macro_name = macro_def_match.group(1).upper()
#         params = macro_def_match.group(2) if macro_def_match.group(2) else ''
#         results['macros_defined'][macro_name] = {
#             'line': line_num,
#             'parameters': params.strip()
#         }
   
#     # Analyze macro calls (excluding system macros)
#     macro_calls = re.findall(r'%([A-Z_][A-Z0-9_]*)\b', line, re.IGNORECASE)
#     system_macros = {
#         'MACRO', 'MEND', 'LET', 'IF', 'THEN', 'ELSE', 'DO', 'END',
#         'EVAL', 'STR', 'QUOTE', 'SCAN', 'SUBSTR', 'INCLUDE', 'ARRAY','MEND','PUT','TO','DOLOOP','SYSFUNC',

#                 # Numeric functions
#         'SUM', 'MEAN', 'MIN', 'MAX', 'COUNT', 'N', 'NMISS',
#         'ROUND', 'CEIL', 'FLOOR', 'INT', 'ABS', 'LOG', 'EXP', 'SQRT',
#         'SIN', 'COS', 'TAN', 'RAND', 'RANUNI', 'NORMAL', 'GAMMA', 'BETA',
        
#         # Character functions
#         'SUBSTR', 'TRIM', 'STRIP', 'LEFT', 'RIGHT', 'LENGTH',
#         'UPCASE', 'LOWCASE', 'PROPCASE', 'COMPRESS', 'TRANSLATE',
#         'INDEX', 'FIND', 'SCAN', 'CATS', 'CATX', 'CAT',
        
#         # Date/Time functions
#         'TODAY', 'DATE', 'DATETIME', 'TIME', 'DATEPART', 'TIMEPART',
#         'YEAR', 'MONTH', 'DAY', 'WEEKDAY', 'MDY', 'YMD',
#         'INTCK', 'INTNX', 'DATDIF', 'JULDATE',
        
#         # Other functions
#         'INPUT', 'PUT', 'COALESCEC', 'COALESCE', 'IFC', 'IFN', 'MISSING'
#     }

#     for macro_name in macro_calls:
#         if macro_name.upper() not in system_macros:
#             results['macros_called'][macro_name.upper()].append(line_num)

def analyze_macros(line, line_num, results, macro_stack):
    """
    Analyze macro definitions and calls with nested macro tracking.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
        macro_stack (list): Stack to track nested macros
    """
    # Analyze macro definitions
    macro_def_match = re.search(r'%MACRO\s+([A-Z_][A-Z0-9_]*)\(([^)]*)\)?\s*;?', line, re.IGNORECASE)
    if macro_def_match:
        macro_name = macro_def_match.group(1).upper()
        params = macro_def_match.group(2) if macro_def_match.group(2) else ''
        
        # Store macro definition with parent information
        results['macros_defined'][macro_name] = {
            'line': line_num,
            'parameters': params.strip(),
            'parent': macro_stack[-1] if macro_stack else None
        }
        
        # Track nested macros
        if macro_stack:
            parent_macro = macro_stack[-1]
            if 'nested_macro_list' not in results:
                results['nested_macro_list'] = {}
            if parent_macro not in results['nested_macro_list']:
                results['nested_macro_list'][parent_macro] = []
            results['nested_macro_list'][parent_macro].append(macro_name)
        
        macro_stack.append(macro_name)  # Push to stack for nested macros
    
    # Analyze macro end statements
    mend_match = re.search(r'%MEND\s*(\w*)?\s*;?', line, re.IGNORECASE)
    if mend_match and macro_stack:
        mend_name = mend_match.group(1)
        if mend_name:
            mend_name = mend_name.upper()
            if mend_name in results['macros_defined']:
                # Update the macro definition with end line
                if 'end_line' not in results['macros_defined'][mend_name]:
                    results['macros_defined'][mend_name]['end_line'] = line_num
                if mend_name in macro_stack:
                    macro_stack.remove(mend_name)
        else:
            # No macro name specified, end the most recent macro
            current_macro = macro_stack.pop()
            if 'end_line' not in results['macros_defined'][current_macro]:
                results['macros_defined'][current_macro]['end_line'] = line_num
   
    # Analyze macro calls (excluding system macros) - YOUR EXISTING CODE
    macro_calls = re.findall(r'%([A-Z_][A-Z0-9_]*)\b', line, re.IGNORECASE)
    system_macros = {
        'MACRO', 'MEND', 'LET', 'IF', 'THEN', 'ELSE', 'DO', 'END',
        'EVAL', 'STR', 'QUOTE', 'SCAN', 'SUBSTR', 'INCLUDE', 'ARRAY','MEND','PUT','TO','DOLOOP','SYSFUNC',

                # Numeric functions
        'SUM', 'MEAN', 'MIN', 'MAX', 'COUNT', 'N', 'NMISS',
        'ROUND', 'CEIL', 'FLOOR', 'INT', 'ABS', 'LOG', 'EXP', 'SQRT',
        'SIN', 'COS', 'TAN', 'RAND', 'RANUNI', 'NORMAL', 'GAMMA', 'BETA',
        
        # Character functions
        'SUBSTR', 'TRIM', 'STRIP', 'LEFT', 'RIGHT', 'LENGTH',
        'UPCASE', 'LOWCASE', 'PROPCASE', 'COMPRESS', 'TRANSLATE',
        'INDEX', 'FIND', 'SCAN', 'CATS', 'CATX', 'CAT',
        
        # Date/Time functions
        'TODAY', 'DATE', 'DATETIME', 'TIME', 'DATEPART', 'TIMEPART',
        'YEAR', 'MONTH', 'DAY', 'WEEKDAY', 'MDY', 'YMD',
        'INTCK', 'INTNX', 'DATDIF', 'JULDATE',
        
        # Other functions
        'INPUT', 'PUT', 'COALESCEC', 'COALESCE', 'IFC', 'IFN', 'MISSING'
    }

    for macro_name in macro_calls:
        if macro_name.upper() not in system_macros:
            results['macros_called'][macro_name.upper()].append(line_num)

def analyze_sql_operations(line, line_num, results):
    """
    Analyze SQL operations within PROC SQL blocks.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
    """
    sql_operations = [
        'SELECT', 'CREATE', 'INSERT', 'UPDATE', 'DELETE', 'ALTER', 'DROP',
        'FROM', 'WHERE', 'GROUP BY', 'HAVING', 'ORDER BY', 'UNION', 'JOIN'
    ]
    
    for operation in sql_operations:
        spaced_pattern = operation.replace(' ', r'\s+')
        word_boundary_pattern = rf'\b{spaced_pattern}\b'
        
        if re.search(word_boundary_pattern, line, re.IGNORECASE):
            operation_key = operation.replace(' ', '_')
            results['sql_statements'][operation_key].append(line_num)


def analyze_control_structures(line, line_num, results):
    """
    Analyze SAS control structures and flow control statements.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
    """
    control_patterns = [
        ('IF_THEN', r'\bIF\s+.*\bTHEN\b'),
        ('DO_LOOP', r'\bDO\b.*(?:TO|WHILE|UNTIL)'),
        ('ARRAY', r'\bARRAY\s+[A-Z_][A-Z0-9_]*'),
        ('FORMAT', r'\bFORMAT\s+'),
        ('LENGTH', r'\bLENGTH\s+'),
        ('LABEL', r'\bLABEL\s+'),
        ('RETAIN', r'\bRETAIN\s+'),
        ('OUTPUT', r'\bOUTPUT\s*;'),
        ('RETURN', r'\bRETURN\s*;'),
        ('DELETE', r'\bDELETE\s*;')
    ]
    
    for structure, pattern in control_patterns:
        if re.search(pattern, line, re.IGNORECASE):
            results['control_structures'][structure].append(line_num)


def analyze_file_operations(line, line_num, results):
    """
    Analyze file operations and library definitions.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
    """
    # Library definitions
    lib_match = re.search(r'LIBNAME\s+([A-Z_][A-Z0-9_]*)', line, re.IGNORECASE)
    if lib_match:
        lib_name = lib_match.group(1).upper()
        results['libraries_defined'][lib_name].append(line_num)
    
    # File operations
    file_ops = [
        ('INFILE', r'\bINFILE\s+'),
        ('FILE', r'\bFILE\s+'),
        ('FILENAME', r'\bFILENAME\s+'),
        ('PUT', r'\bPUT\s+'),
        ('INPUT', r'\bINPUT\s+')
    ]
    
    for op_name, pattern in file_ops:
        if re.search(pattern, line, re.IGNORECASE):
            results['file_operations'][op_name].append(line_num)


def analyze_variables(line, line_num, results):
    """
    Analyze variable operations and data manipulation statements.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
    """
    var_ops = [
        ('KEEP', r'KEEP\s+([A-Z_][A-Z0-9_]*(?:\s+[A-Z_][A-Z0-9_]*)*?)(?:\s|;|$)'),
        ('DROP', r'DROP\s+([A-Z_][A-Z0-9_]*(?:\s+[A-Z_][A-Z0-9_]*)*?)(?:\s|;|$)'),
        ('VAR', r'VAR\s+([A-Z_][A-Z0-9_]*(?:\s+[A-Z_][A-Z0-9_]*)*?)(?:\s|;|$)'),
        ('BY', r'BY\s+([A-Z_][A-Z0-9_]*(?:\s+[A-Z_][A-Z0-9_]*)*?)(?:\s|;|$)'),
        ('CLASS', r'CLASS\s+([A-Z_][A-Z0-9_]*(?:\s+[A-Z_][A-Z0-9_]*)*?)(?:\s|;|$)')
    ]
    
    for op_name, pattern in var_ops:
        matches = re.findall(pattern, line, re.IGNORECASE)
        for match in matches:
            variables = match.split()
            for var in variables:
                results['variables_used'][f"{op_name}_{var.upper()}"].append(line_num)


def check_block_start(line, line_num, results, state):
    """
    Detect the start of SAS code blocks (DATA, PROC, MACRO).
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
        state (dict): Current analysis state
    """
    # DATA step blocks
    data_match = re.match(r'^\s*DATA\s+([A-Z_][A-Z0-9_.]*(?:\s+[A-Z_][A-Z0-9_.]*)*)', line, re.IGNORECASE)
    if data_match:
        datasets = data_match.group(1).split()
        block_info = {
            'type': 'DATA',
            'name': f"DATA {' '.join(datasets)}",
            'start_line': line_num,
            'end_line': None,
            'datasets': datasets
        }
        state['current_blocks'].append(block_info)
        return
    
    # PROC blocks
    proc_match = re.match(r'^\s*PROC\s+([A-Z]+)(?:\s+DATA\s*=\s*([A-Z_][A-Z0-9_.]*))?\s*;?', line, re.IGNORECASE)
    if proc_match:
        proc_name = proc_match.group(1).upper()
        dataset = proc_match.group(2).upper() if proc_match.group(2) else 'UNKNOWN'
        block_info = {
            'type': 'PROC',
            'name': f"PROC {proc_name}",
            'start_line': line_num,
            'end_line': None,
            'proc_name': proc_name,
            'dataset': dataset
        }
        state['current_blocks'].append(block_info)
        return
    
    # MACRO blocks
    macro_match = re.match(r'^\s*%MACRO\s+([A-Z_][A-Z0-9_]*)\(([^)]*)\)?\s*;?', line, re.IGNORECASE)
    if macro_match:
        macro_name = macro_match.group(1).upper()
        params = macro_match.group(2) if macro_match.group(2) else ''
        block_info = {
            'type': 'MACRO',
            'name': f"%MACRO {macro_name}",
            'start_line': line_num,
            'end_line': None,
            'macro_name': macro_name,
            'parameters': params.strip()
        }
        state['current_blocks'].append(block_info)


def check_block_end(line, line_num, results, state):
    """
    Detect the end of SAS code blocks.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        results (dict): Results dictionary to update
        state (dict): Current analysis state
    """
    if re.search(r'\b(?:RUN|QUIT)\s*;', line, re.IGNORECASE) or '%MEND' in line.upper():
        if state['current_blocks']:
            block = state['current_blocks'].pop()
            block['end_line'] = line_num
            results['function_blocks'].append(block)


def finalize_open_blocks(total_lines, results, state):
    """
    Close any remaining open blocks at the end of file analysis.
    
    Args:
        total_lines (int): Total number of lines in the file
        results (dict): Results dictionary to update
        state (dict): Current analysis state
    """
    while state['current_blocks']:
        block = state['current_blocks'].pop()
        block['end_line'] = total_lines
        results['function_blocks'].append(block)


def extract_detailed_proc_info(results):
    """
    Extract detailed information for PROC IMPORT/EXPORT and DATA steps.
    
    Args:
        results (dict): Results dictionary to update
    """
    # Extract PROC IMPORT details
    for block in results['function_blocks']:
        if block.get('proc_name') == 'IMPORT':
            options = extract_proc_options(results['line_analysis'], 
                                         block['start_line'], block['end_line'], 'IMPORT')
            results['proc_import_details'].append({
                'name': 'PROC IMPORT',
                'line_number': block['start_line'],
                'out': options.get('OUT', 'Not specified'),
                'dbms': options.get('DBMS', 'Not specified'),
                'datafile': options.get('DATAFILE', 'Not specified')
            })
    
    # Extract PROC EXPORT details
    for block in results['function_blocks']:
        if block.get('proc_name') == 'EXPORT':
            options = extract_proc_options(results['line_analysis'], 
                                         block['start_line'], block['end_line'], 'EXPORT')
            results['proc_export_details'].append({
                'name': 'PROC EXPORT',
                'line_number': block['start_line'],
                'data': options.get('DATA', 'Not specified'),
                'dbms': options.get('DBMS', 'Not specified'),
                'outfile': options.get('OUTFILE', 'Not specified')
            })
    
    # Extract DATA step details
    for block in results['function_blocks']:
        if block.get('type') == 'DATA':
            dataset_name = ' '.join(block.get('datasets', ['Unknown']))
            dataset_type = determine_dataset_type(block.get('datasets', [''])[0])
            operations = extract_data_step_operations(results['line_analysis'], 
                                                    block['start_line'], block['end_line'])
            size = block['end_line'] - block['start_line'] + 1 if block['end_line'] else 'Unknown'
            
            results['data_step_details'].append({
                'name': dataset_name,
                'type': dataset_type,
                'line_number': block['start_line'],
                'size': size,
                'operations_used': ', '.join(operations) if operations else 'None detected'
            })


def calculate_complexity_metrics(results):
    """
    Calculate code complexity metrics based on analysis results.
    
    Args:
        results (dict): Results dictionary to update
    """
    results['code_complexity'] = {
        'total_functions': len(results['function_blocks']),
        'total_datasets': len(results['datasets_created']) + len(results['datasets_used']),
        'total_procedures': len(results['procedures_used']),
        'total_macros': len(results['macros_defined']),
        'total_includes': len(results['include_files']),
        'total_sys_functions': len(results['system_functions']),
        'total_lines': len(results['line_analysis'])
    }


def analyze_single_line(line, line_num, total_lines, source_file, results, state):
    """
    Perform comprehensive analysis of a single line of SAS code.
    
    Args:
        line (str): Code line to analyze
        line_num (int): Line number
        total_lines (int): Total lines in file
        source_file (str): Source file name
        results (dict): Results dictionary to update
        state (dict): Current analysis state
    """
    original_line = line.strip()
    cleaned_line = clean_line(line).strip()
    
    # Store line analysis (include blank and comment lines)
    results['line_analysis'][line_num] = {
        'original': original_line,
        'cleaned': cleaned_line,
        'type': classify_line(cleaned_line),
        'source_file': source_file
    }
    
    # Skip further analysis for empty lines
    if not cleaned_line:
        return
    
    # Perform all analysis functions
    check_block_start(cleaned_line, line_num, results, state)
    analyze_data_operations(cleaned_line, line_num, results)
    analyze_procedures(cleaned_line, line_num, results)
    analyze_macros(cleaned_line, line_num, results,state['macro_stack'])
    analyze_sql_operations(cleaned_line, line_num, results)
    analyze_control_structures(cleaned_line, line_num, results)
    analyze_file_operations(cleaned_line, line_num, results)
    analyze_variables(cleaned_line, line_num, results)
    analyze_include_files(cleaned_line, line_num, results)
    analyze_system_functions(cleaned_line, line_num, results)
    analyze_snowflake_references(line, line_num, results, state)
    check_block_end(cleaned_line, line_num, results, state)


def analyze_lines(lines, source_file=None):
    """
    Analyze all lines in a SAS code file.
    
    Args:
        lines (list): List of code lines
        source_file (str): Source file name
        
    Returns:
        dict: Analysis results
    """
    results, state = initialize_analysis()
    total_lines = len(lines)
    
    for line_num, line in enumerate(lines, 1):
        try:
            analyze_single_line(line, line_num, total_lines, source_file, results, state)
        except Exception as e:
            print(f"Warning: Error analyzing line {line_num}: {e}")
            continue
    
    finalize_open_blocks(total_lines, results, state)
    extract_detailed_proc_info(results)
    calculate_complexity_metrics(results)
    return results


def analyze_file(file_path):
    """
    Load and analyze a SAS code file.
    
    Args:
        file_path (str): Path to SAS file
        
    Returns:
        dict: Analysis results or None if error
    """
    if not os.path.exists(file_path):
        print(f"❌ Error: File '{file_path}' not found.")
        return None
    
    try:
        # Try different encodings
        encodings = ['utf-8', 'latin-1', 'cp1252']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                    lines = f.readlines()
                print(f"✅ Successfully loaded file: {file_path} ({len(lines)} lines) with encoding: {encoding}")
                return analyze_lines(lines, file_path)
            except UnicodeDecodeError:
                continue
        
        print(f"❌ Could not decode file with any supported encoding")
        return None
        
    except Exception as e:
        print(f"❌ Error reading file {file_path}: {e}")
        return None


# =====================================================
# PROFESSIONAL EXCEL STYLING FUNCTIONS
# =====================================================

def apply_cell_style(cell, style_name):
    """
    Apply a predefined style to a cell.
    
    Args:
        cell: Excel cell object
        style_name (str): Name of style to apply
    """
    style = EXCEL_STYLES.get(style_name, EXCEL_STYLES['data_cell'])
    cell.font = style['font']
    cell.alignment = style['alignment']
    if 'fill' in style:
        cell.fill = style['fill']
    cell.border = BORDER_THIN


def create_section_header(ws, row, col_span, title, style='section_header'):
    """
    Create a styled section header.
    
    Args:
        ws: Worksheet object
        row (int): Row number
        col_span (int): Number of columns to span
        title (str): Header title
        style (str): Style name
        
    Returns:
        int: Next row number
    """
    if col_span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    
    header_cell = ws.cell(row=row, column=1, value=title)
    apply_cell_style(header_cell, style)
    header_cell.border = BORDER_THICK
    
    return row + 1


def create_table_with_data(ws, start_row, headers, data, title=None):
    """
    Create a professionally styled table with data.
    
    Args:
        ws: Worksheet object
        start_row (int): Starting row
        headers (list): Table headers
        data (list): Table data
        title (str): Optional table title
        
    Returns:
        int: Next row number after table
    """
    current_row = start_row
    # Add title if provided
    if title:
        current_row = create_section_header(ws, current_row, len(headers), title)
        current_row += 1
    
    # Add table headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        apply_cell_style(cell, 'table_header')
    current_row += 1
    
    # Add data rows with alternating colors
    for row_idx, row_data in enumerate(data):
        style_name = 'data_cell' if row_idx % 2 == 0 else 'data_cell_alt'
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=str(value))
            apply_cell_style(cell, style_name)
        current_row += 1
    
    return current_row + 1


def auto_adjust_columns(ws):
    """
    Auto-adjust column widths based on content.
    
    Args:
        ws: Worksheet object
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set reasonable column width
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)


def create_professional_excel_report(results, output_file):
    """
    Create a professional Excel report with attractive styling.
    
    Args:
        results (dict): Analysis results
        output_file (str): Output Excel file path
        
    Returns:
        str: Output file path or None if error
    """
    try:
        # Handle file permission issues
        if os.path.exists(output_file):
            try:
                os.remove(output_file)
            except PermissionError:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_name = os.path.splitext(output_file)[0]
                output_file = f"{base_name}_{timestamp}.xlsx"
                print(f"⚠️ Original file is open. Using new filename: {output_file}")
        
        # Create workbook with initial sheets
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            pd.DataFrame().to_excel(writer, sheet_name='SAS Analysis', index=False)
            pd.DataFrame().to_excel(writer, sheet_name='Function Blocks', index=False)
            pd.DataFrame().to_excel(writer, sheet_name='Queries Analysis', index=False)

        # Load workbook for styling
        wb = load_workbook(output_file)

        # =====================================================
        # MAIN ANALYSIS SHEET
        # =====================================================
        ws_main = wb['SAS Analysis']
        ws_main.delete_rows(1, ws_main.max_row)
        current_row = 1

        # Main title
        current_row = create_section_header(ws_main, current_row, 5, 
                                          'SAS CODE ANALYSIS REPORT', 'main_header')
        current_row += 1
        
        # Summary Metrics
        summary_data = []
        for key, value in results['code_complexity'].items():
            summary_data.append([key.replace('_', ' ').title(), value])
        
        current_row = create_table_with_data(
            ws_main, current_row, 
            ['📊 Metric', '📈 Value'], 
            summary_data, 
            'CODE COMPLEXITY SUMMARY'
        )

        # MACROS DEFINED
        # 'CLAIMS': {'line': 362, 'parameters': 'MONTHEND', 'parent': None, 'end_line': 455}
        if results['macros_defined']:
            macro_data = []
            for macro_name, info in results['macros_defined'].items():
                line_range = str(info['line']) + '→' + str(info['end_line'])
                macro_data.append([
                    macro_name,
                    line_range,
                    info['parameters'] if info['parameters'] else 'None',
                    info['parent'] if info['parent'] else '-',

                ])

            current_row = create_table_with_data(
                ws_main, current_row,
                ['📊 Macro Name', '📍 Line Ranges', '🔢 Parameters','📌 Parent'],
                macro_data,
                'MACROS DEFINED'
            )

    # print(results['macros_called'])

        # MACROS CALLED
        if results['macros_called']:
            macro_data = []
            for macro_name, location in results['macros_called'].items():
                location_str =  ",".join(map(str,location))
                macro_data.append([
                    macro_name,
                    location_str
                ])

            current_row = create_table_with_data(
                ws_main, current_row,
                ['📊 Macro Name', '📍 Called Line Numbers'],
                macro_data,
                'MACROS CALLED'
            )

        # INCLUDE FILES
        if results['include_files']:
            include_data = []
            for include_name, lines in results['include_files'].items():
                include_data.append([
                    include_name,
                    lines[0] if lines else 'Unknown'
                ])

            current_row = create_table_with_data(
                ws_main, current_row,
                ['📊 Include Name', '📍 Line Number'],
                include_data,
                'INCLUDE FILES'
            )

        # SYSTEM FUNCTIONS
        if results['system_functions']:
            func_data = []
            for func_name, lines in results['system_functions'].items():
                # Limit displayed line numbers to avoid overly long cells
                line_display = ', '.join(map(str, lines[:10]))
                if len(lines) > 100:
                    line_display += f" ... (and {len(lines)-100} more)"
                
                func_data.append([
                    func_name,
                    len(lines),
                    line_display
                ])
            
            current_row = create_table_with_data(
                ws_main, current_row,
                ['⚙️ Function Name', '🔢 Count', '📍 Line Numbers'],
                func_data,
                'SYSTEM FUNCTIONS'
            )
        
        auto_adjust_columns(ws_main)

        # =====================================================
        # FUNCTION BLOCKS SHEET
        # =====================================================
        ws_fb = wb['Function Blocks']
        ws_fb.delete_rows(1, ws_fb.max_row)
        current_row = 1

        # Main title
        current_row = create_section_header(ws_fb, current_row, 5, 
                                          'FUNCTION BLOCKS - DATA & PROC', 'main_header')
        current_row += 1

        # PROC IMPORT DETAILS
        if results['proc_import_details']:
            import_data = []
            for proc_info in results['proc_import_details']:
                import_data.append([
                    proc_info['name'],
                    proc_info['line_number'],
                    proc_info['out'],
                    proc_info['dbms'],
                    proc_info['datafile']
                ])
            
            current_row = create_table_with_data(
                ws_fb, current_row,
                ['🔧 Name (PROC IMPORT)', '📍 Line Number', '📤 OUT', '💾 DBMS', '📁 DATAFILE'],
                import_data,
                'PROC IMPORT DETAILS'
            )

        # PROC EXPORT DETAILS
        if results['proc_export_details']:
            export_data = []
            for proc_info in results['proc_export_details']:
                export_data.append([
                    proc_info['name'],
                    proc_info['line_number'],
                    proc_info['data'],
                    proc_info['dbms'],
                    proc_info['outfile']
                ])
            
            current_row = create_table_with_data(
                ws_fb, current_row,
                ['🔧 Name (PROC EXPORT)', '📍 Line Number', '📥 DATA', '💾 DBMS', '📁 OUTFILE'],
                export_data,
                'PROC EXPORT DETAILS'
            )

        # REMAINING PROC STEPS - Group by proc_name and list line ranges with count
        proc_blocks = [block for block in results['function_blocks'] 
                      if block.get('name', '').startswith('PROC') 
                      and block.get('proc_name') not in ['IMPORT', 'EXPORT']]
        
        if proc_blocks:
            proc_grouped = defaultdict(list)
            for block in proc_blocks:
                line_range = f"{block.get('start_line', 'N/A')} → {block.get('end_line', 'N/A')}"
                proc_grouped[block.get('proc_name', 'UNKNOWN')].append(line_range)
            
            proc_step_data = []
            for proc_name, ranges in proc_grouped.items():
                joined_ranges = ' , '.join(ranges)
                proc_step_data.append([
                    f"PROC {proc_name}",
                    len(ranges),
                    joined_ranges
                ])
  
            current_row = create_table_with_data(
                ws_fb, current_row,
                ['🔧 Proc Step', '🔢 Count', '📍 Line Ranges'],
                proc_step_data,
                'OTHER PROC STEPS'
            )

        # # DATA STEPS DETAILS
        # if results['data_step_details']:
        #     data_step_data = []
        #     for data_info in results['data_step_details']:
        #         data_step_data.append([
        #             data_info['name'],
        #             data_info['type'],
        #             data_info['line_number'],
        #             data_info['size'],
        #             data_info['operations_used']
        #         ])
            
        #     current_row = create_table_with_data(
        #         ws_fb, current_row,
        #         ['📊 Name', '📌 Type (Temp/Permanent)', '📍 Line Number', '📏 Size', '⚙️ Operations Used'],
        #         data_step_data,
        #         'DATA STEPS DETAILS'
        #     )

        if results['data_step_details']:
            data_grouped = defaultdict(list)
            for data_info in results['data_step_details']:
                # Use dataset name as grouping key
                data_grouped[data_info['name']].append(data_info)

            data_step_data = []
            for dataset_name, entries in data_grouped.items():
                lines = [entry['line_number'] for entry in entries]
                size = [entry['size'] for entry in entries]
                new = []
                for line_num, size_val in zip(lines, size):
                    new.append(f"{line_num} → {line_num + size_val - 1}")

                line_range = ', '.join(map(str, new[:100]))

                count_blocks = len(entries)
                # Operations: collect unique operations from all grouped entries
                operations_set = set()
                for entry in entries:
                    if entry['operations_used'] and entry['operations_used'] != 'None detected':
                        operations_set.update(op.strip() for op in entry['operations_used'].split(','))

                operations_str = ', '.join(sorted(operations_set)) if operations_set else 'None detected'

                # Determine type — heuristic: if dataset starts with "WORK." or no libref, Temporary else Permanent
                type_label = 'Temporary' if dataset_name.upper().startswith('WORK.') or '.' not in dataset_name else 'Permanent'
                dataset_name = 'DATA ' + dataset_name                

                data_step_data.append([
                    dataset_name,
                    type_label,
                    count_blocks,
                    operations_str,
                    line_range
                ])

            current_row = create_table_with_data(
                ws_fb, current_row,
                ['📊 Name', '📌 Type', '🔢 Count', '⚙️ Operations Used', '📍 Line Range(s)'],
                data_step_data,
                'DATA STEPS DETAILS'
            )


        auto_adjust_columns(ws_fb)

        # =====================================================
        # QUERIES ANALYSIS SHEET
        # =====================================================
        ws_snow = wb['Queries Analysis']
        ws_snow.delete_rows(1, ws_snow.max_row)
        current_row = 1
        
        # Main title
        current_row = create_section_header(ws_snow, current_row, 7, 
                                          'QUERIES ANALYSIS', 'main_header')
        current_row += 1
        
        # Separate connection queries from other queries
        connection_queries = []
        regular_queries = []
        
        for idx, query_info in enumerate(results['snowflake_queries'], 1):
            query_lower = query_info['query'].lower()
            
            line_range = f"{query_info['start_line']} → {query_info['end_line']}"
            tables_ref = ", ".join(query_info['tables_referenced']) if query_info['tables_referenced'] else 'None detected'
            
            # Determine query type
            query_upper = query_info['query'].upper()
            if 'SELECT' in query_upper:
                query_type = '📊 SELECT'
            elif 'INSERT' in query_upper:
                query_type = '➕ INSERT'
            elif 'UPDATE' in query_upper:
                query_type = '✏️ UPDATE'
            elif 'CREATE' in query_upper:
                query_type = '🏗️ CREATE'
            elif 'CONNECT' in query_upper:
                query_type = '🔌 CONNECT'
            else:
                query_type = '❓ OTHER'
            
            query_data = [
                f"Query {idx}",
                line_range,
                query_type,
                tables_ref,
                query_info['created_table_name'],
                query_info['created_table_type'],
                query_info['query']  # Full query
            ]
            
            if query_lower.startswith("connect"):
                connection_queries.append(query_data)
            else:
                regular_queries.append(query_data)

        # Display connection queries first
        if connection_queries:
            current_row = create_table_with_data(
                ws_snow, current_row,
                ['🔍 Query ID', '📍 Line Range', '🔧 Type', '📊 Tables Used', '🗃️ Created Table Name', '📌 Created Table Type', '📝 Full Query'],
                connection_queries,
                'CONNECTION QUERIES'
            )

        # Display regular SQL queries
        if regular_queries:
            current_row = create_table_with_data(
                ws_snow, current_row,
                ['🔍 Query ID', '📍 Line Range', '🔧 Type', '📊 Tables Used', '🗃️ Created Table Name', '📌 Created Table Type', '📝 Full Query'],
                regular_queries,
                'SQL QUERIES ANALYSIS'
            )

        # Show message if no queries found
        if not connection_queries and not regular_queries:
            no_queries_cell = ws_snow.cell(row=current_row, column=1, 
                                        value='ℹ️ No SQL queries detected in the code')
            no_queries_cell.font = Font(name='Calibri', size=12, italic=True, color='888888')
        
        auto_adjust_columns(ws_snow)
        
        # Set the main sheet as active
        wb.active = ws_main
        
        # Save the workbook
        wb.save(output_file)
        print(f"✅ Professional Excel report saved at: {output_file}")
        
        return output_file
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        import traceback
        traceback.print_exc()
        return None


# =====================================================
# MAIN PROGRAM ENTRY POINT
# =====================================================

def analyze_sas_file(input_file_path, excel_output='sas_analysis_report.xlsx'):
    """
    Main function to analyze a SAS file and generate professional Excel report.
    
    Args:
        input_file_path (str): Path to input SAS file
        excel_output (str): Path for output Excel file
        
    Returns:
        dict: Analysis results or None if error
    """
    print("🚀 Starting SAS Code Analysis...")
    print("=" * 70)
    
    results = analyze_file(input_file_path)
    if not results:
        print("❌ Analysis failed")
        return None
    
    print("📊 Analysis completed. Generating professional Excel report...")
    output_file = create_professional_excel_report(results, excel_output)
    
    if output_file:
        print(f"✅ Professional Excel report generated: {output_file}")
        print("🎨 Report features:")
        print("   • 🎨 Professional color scheme with blues and accent colors")
        print("   • 📊 Styled headers with gradients and borders")
        print("   • 🔄 Alternating row colors for easy reading")
        print("   • 📱 Auto-sized columns for optimal viewing")
        print("   • ❄️ Dedicated Snowflake analysis with table structure breakdown")
        print("   • 🔍 Query type classification with emojis")
        print("   • 📈 Summary metrics dashboard")
        print("   • 📝 Full SQL queries displayed (not truncated)")
        print("   • 🗃️ Dataset creation type detection (Temporary/Permanent)")
        print("   • 📋 Detailed PROC IMPORT/EXPORT analysis with parameters")
        print("   • 🔢 Grouped PROC step analysis with counts and line ranges")
        print("   • ⚙️ DATA step operations tracking")
        print("   • 🏗️ Enhanced query analysis with created table info")
    
    return results


if __name__ == "__main__":
    print("🔍 Professional SAS Code Analyzer")
    print("🎨 With Beautiful Excel Styling")
    print("=" * 60)
    
    # Default input file
    input_file = 'a.c'
    
    if not os.path.exists(input_file):
        print(f"❌ File not found: {input_file}")
        print("Please ensure the SAS file exists in the current directory.")
        exit(1)
        
    # Fixed filename generation
    base_name = os.path.splitext(input_file)[0]  # Gets 'test' from 'test.sas'
    output_xlsx = f"{base_name}_sas_analysis_report.xlsx"

    if not output_xlsx:
        output_xlsx = "sas_analysis_report.xlsx"
    
    try:
        # Perform analysis
        results = analyze_sas_file(input_file, excel_output=output_xlsx)
        
        if results:
            print("\n🎉 Analysis Complete!")
            print("📊 Your beautiful, professional SAS analysis report is ready!")
        else:
            print("\n❌ Analysis failed. Please check the input file and try again.")
            
    except KeyboardInterrupt:
        print("\n⚠️ Analysis interrupted by user.")
    except Exception as e:
        print(f"\n❌ Unexpected error during analysis: {e}")
        import traceback
        traceback.print_exc()
