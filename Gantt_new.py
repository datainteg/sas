import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ==== USER SETTINGS ====
INPUT_FILE = "Project_Scripts_Conversion_Tracking.xlsx"
SHEET_NAME = "Sheet1"
OUTPUT_FILE = "UAT_Flag_Reordered_Gantt.xlsx"
# ===========

def create_uat_flags_by_uniqueid(df):
    df = df.copy()
    df['Analysis Start Date'] = pd.to_datetime(df['Analysis Start Date'], errors='coerce').dt.strftime('%Y-%m-%d')
    df['UAT End Date'] = pd.to_datetime(df['UAT End Date'], errors='coerce').dt.strftime('%Y-%m-%d')
    df['Analysis Start Date'] = df['Analysis Start Date'].fillna('')
    df['UAT End Date'] = df['UAT End Date'].fillna('')
    df['UniqueID'] = df['Analysis Start Date'] + '_' + df['UAT End Date']
    
    unique_uids = sorted([uid for uid in df['UniqueID'].unique() if uid != '_'])
    uid_to_uat_map = {uid: f"UAT-{i + 1}" for i, uid in enumerate(unique_uids)}
    df['UAT_Flag'] = df['UniqueID'].map(uid_to_uat_map).fillna('No-UAT')
    
    return df, uid_to_uat_map

def build_uat_flag_reordered_gantt():
    try:
        df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
        print(f"✅ Successfully loaded {len(df)} rows from {INPUT_FILE}")
    except Exception as e:
        raise Exception(f"❌ Error reading Excel file: {e}")

    # Clean data
    df['Business Function Owner'] = df['Business Function Owner'].fillna('Unknown')
    df['Path'] = df['Path'].fillna('Unknown') 
    df['Script Name'] = df['Script Name'].fillna('Unknown')
    df['Priority'] = df['Priority'].fillna(1)

    # Create UAT flags
    df_with_uat, uat_mapping = create_uat_flags_by_uniqueid(df)

    # **PROFESSIONAL STANDARD COLORS** based on research
    phase_definitions = [
        ("Analysis", "Analysis Start Date", "Analysis End Date", "FF4F81BD"),      # Professional Blue
        ("Conversion", "Conversion Start Date", "Conversion End Date", "FF9BBB59"), # Professional Olive Green
        ("Execution", "Execution Start Date", "Execution End Date", "FFED7D31"),    # Professional Orange
        ("Recon", "Recon Start Date", "Recon End Date", "FF8C8C8C"),                # Professional Gray
        ("Issue Fix", "Issue Fix Start Date", "Issue Fix End Date", "FFBF7F00"),    # Professional Amber
        ("UAT", "UAT Start Date", "UAT End Date", "FF70AD47")                       # **GREEN** as requested
    ]

    available_phases = [(name, start, end, color) for name, start, end, color in phase_definitions 
                       if start in df.columns and end in df.columns]

    # Create task structure
    all_tasks = []
    for idx, row in df_with_uat.iterrows():
        uat_flag = row['UAT_Flag']
        script_phases = []
        
        for phase_name, start_col, end_col, color in available_phases:
            s_val = df.iloc[idx].get(start_col)
            e_val = df.iloc[idx].get(end_col)
            
            if (pd.notna(s_val) and pd.notna(e_val) and str(s_val) != "########" and str(e_val) != "########"):
                try:
                    start_dt = pd.to_datetime(s_val, errors='coerce')
                    end_dt = pd.to_datetime(e_val, errors='coerce')
                    
                    if pd.notna(start_dt) and pd.notna(end_dt):
                        script_phases.append({
                            "Phase": phase_name,
                            "Start": start_dt.date(),
                            "End": end_dt.date(),
                            "Color": color
                        })
                except:
                    continue
        
        if script_phases:
            all_tasks.append({
                "UAT_Flag": uat_flag,
                "Business Function Owner": str(row['Business Function Owner']),
                "Path": str(row['Path']),
                "Script Name": str(row['Script Name']),
                "Priority": row['Priority'],
                "Phases": script_phases
            })

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "UAT_Reordered_Gantt"

    # Title
    ws.merge_cells("A1:R1")
    title = ws.cell(1, 1, "GANTT CHART — UAT FLAG AFTER PHASE COLUMN")
    title.font = Font(name="Calibri", size=16, bold=True, color="FFFFFFFF")
    title.fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")
    title.alignment = Alignment(horizontal="center")

    # **KEY CHANGE**: Headers with UAT Flag AFTER Phase
    headers = ["Business Function Owner", "Path", "Script Name", "Priority", "Phase", "UAT Flag", "Start", "End"]
    
    for i, h in enumerate(headers, 1):
        cell = ws.cell(3, i, h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")
        ws.column_dimensions[get_column_letter(i)].width = 25 if h in ["Business Function Owner", "Path", "Script Name"] else 12

    # Calculate timeline
    all_starts = [phase["Start"] for task in all_tasks for phase in task["Phases"]]
    all_ends = [phase["End"] for task in all_tasks for phase in task["Phases"]]
    min_date = min(all_starts)
    max_date = max(all_ends)
    total_days = (max_date - min_date).days + 1

    # Timeline headers with professional weekend highlighting
    timeline_start_col = len(headers) + 1
    for d in range(total_days):
        dt = min_date + timedelta(days=d)
        col = timeline_start_col + d
        cell = ws.cell(3, col, dt.strftime("%d\n%b"))
        cell.font = Font(bold=True, color="FFFFFFFF", size=9)
        
        # Professional weekend highlighting
        if dt.weekday() >= 5:
            cell.fill = PatternFill(start_color="FFDC143C", end_color="FFDC143C", fill_type="solid")  # Crimson
        else:
            cell.fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")  # Professional Blue
            
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 4

    # **KEY CHANGE**: Write data with UAT Flag in column 6 (after Phase)
    thin = Side(style="thin", color="FF000000")
    thick = Side(style="thick", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    current_row = 4
    
    for task in all_tasks:
        phases = task["Phases"]
        num_phases = len(phases)
        script_start_row = current_row
        script_end_row = current_row + num_phases - 1

        # Merge script info columns (1-4: Business Owner, Path, Script Name, Priority)
        script_merge_columns = {
            1: task["Business Function Owner"], 
            2: task["Path"],
            3: task["Script Name"],
            4: task["Priority"]
        }

        for col_idx, value in script_merge_columns.items():
            if num_phases > 1:
                ws.merge_cells(start_row=script_start_row, start_column=col_idx, end_row=script_end_row, end_column=col_idx)
            cell = ws.cell(script_start_row, col_idx, value)
            cell.alignment = Alignment(vertical="center", horizontal="left" if col_idx in [1,2,3] else "center", wrap_text=True)

        # **REORDERED**: Write Phase (column 5), UAT Flag (column 6), then Start/End (columns 7,8)
        for phase in phases:
            ws.cell(current_row, 5, phase["Phase"])          # Phase column
            ws.cell(current_row, 6, task["UAT_Flag"])        # **UAT Flag AFTER Phase**
            ws.cell(current_row, 7, phase["Start"])          # Start column
            ws.cell(current_row, 8, phase["End"])            # End column

            # Professional Gantt timeline bars with phase names merged horizontally
            phase_start_col = timeline_start_col + (phase["Start"] - min_date).days
            phase_end_col = timeline_start_col + (phase["End"] - min_date).days
            
            # Merge phase name horizontally across duration
            if phase_end_col > phase_start_col:
                ws.merge_cells(start_row=current_row, start_column=phase_start_col, 
                             end_row=current_row, end_column=phase_end_col)
            
            # Write phase name in merged cell with professional color
            merged_cell = ws.cell(current_row, phase_start_col, phase["Phase"])
            merged_cell.fill = PatternFill(start_color=phase["Color"], end_color=phase["Color"], fill_type="solid")
            merged_cell.font = Font(bold=True, color="FFFFFFFF", size=9)
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1

        # Professional script block borders
        for r in range(script_start_row, script_end_row + 1):
            for c in range(1, timeline_start_col + total_days):
                cell = ws.cell(r, c)
                
                top_border = thick if r == script_start_row else thin
                bottom_border = thick if r == script_end_row else thin
                left_border = thick if c == 1 else thin
                right_border = thick if c == timeline_start_col + total_days - 1 else thin
                
                cell.border = Border(top=top_border, bottom=bottom_border, 
                                   left=left_border, right=right_border)

        # Professional separator row
        for c in range(1, timeline_start_col + total_days):
            separator_cell = ws.cell(current_row, c)
            separator_cell.fill = PatternFill(start_color="FFF0F0F0", end_color="FFF0F0F0", fill_type="solid")
        current_row += 1

    # Apply professional finishing touches
    max_row = ws.max_row
    max_col = timeline_start_col + total_days - 1

    # Today marker
    today = datetime.now().date()
    if min_date <= today <= max_date:
        today_col = timeline_start_col + (today - min_date).days
        for r in range(3, max_row + 1):
            ws.cell(r, today_col).border = Border(left=Side(style="thick", color="FFFF0000"))

    # Excel Table for filtering
    table_ref = f"A3:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="UATReorderedTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    ws.freeze_panes = "F4"
    wb.save(OUTPUT_FILE)
    
    print(f"🎉 SUCCESS: UAT Flag reordered Gantt chart created - {OUTPUT_FILE}")
    print("✅ Changes Applied:")
    print("   • UAT Flag moved to column 6 (AFTER Phase column)")
    print("   • Professional standard colors implemented")
    print("   • UAT phase kept GREEN as requested")
    print("   • Phase names merged horizontally across durations")
    print("   • Professional borders and styling")
    
    return OUTPUT_FILE

if __name__ == "__main__":
    try:
        build_uat_flag_reordered_gantt()
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
