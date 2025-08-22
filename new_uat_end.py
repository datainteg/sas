import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ==== USER SETTINGS ====
INPUT_FILE = "Project_Scripts_Conversion_Tracking.xlsx"
SHEET_NAME = "Sheet1"
OUTPUT_FILE = "Gantt_With_UAT_End_Date.xlsx"
# ===========

def create_uat_flags_by_uniqueid(df):
    df = df.copy()
    df['Analysis Start Date'] = pd.to_datetime(df['Analysis Start Date'], errors='coerce').dt.strftime('%Y-%m-%d')
    df['UAT End Date'] = pd.to_datetime(df['UAT End Date'], errors='coerce').dt.strftime('%Y-%m-%d')
    df['Analysis Start Date'] = df['Analysis Start Date'].fillna('')
    df['UAT End Date'] = df['UAT End Date'].fillna('')
    df['UniqueID'] = df['Analysis Start Date'] + '_' + df['UAT End Date']

    unique_uids = sorted([uid for uid in df['UniqueID'].unique() if uid != '_'])
    uid_to_uat_map = {uid: f"UAT-{i + 1}" for i, uid in enumerate(unique_uids)}
    df['UAT_Flag'] = df['UniqueID'].map(uid_to_uat_map).fillna('No-UAT')

    return df, uid_to_uat_map

def build_gantt_with_uat_end():
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
    df['Business Function Owner'] = df['Business Function Owner'].fillna('Unknown')
    df['Path'] = df['Path'].fillna('Unknown')
    df['Script Name'] = df['Script Name'].fillna('Unknown')
    df['Priority'] = df['Priority'].fillna(1)

    df_with_uat, uat_mapping = create_uat_flags_by_uniqueid(df)

    phase_definitions = [
        ("Analysis", "Analysis Start Date", "Analysis End Date", "FF4F81BD"),
        ("Conversion", "Conversion Start Date", "Conversion End Date", "FF9BBB59"),
        ("Execution", "Execution Start Date", "Execution End Date", "FFED7D31"),
        ("Recon", "Recon Start Date", "Recon End Date", "FF8C8C8C"),
        ("Issue Fix", "Issue Fix Start Date", "Issue Fix End Date", "FFBF7F00"),
        ("UAT", "UAT Start Date", "UAT End Date", "FF70AD47"),
    ]

    available_phases = [(name, start, end, color) for name, start, end, color in phase_definitions
                        if start in df.columns and end in df.columns]

    all_tasks = []
    for idx, row in df_with_uat.iterrows():
        uat_flag = row['UAT_Flag']
        uat_end_date = row['UAT End Date'] if 'UAT End Date' in row and pd.notna(row['UAT End Date']) else ''
        script_phases = []

        for phase_name, start_col, end_col, color in available_phases:
            s_val = df.iloc[idx].get(start_col)
            e_val = df.iloc[idx].get(end_col)

            if (pd.notna(s_val) and pd.notna(e_val) and str(s_val) != "########" and str(e_val) != "########"):
                try:
                    start_dt = pd.to_datetime(s_val, errors='coerce')
                    end_dt = pd.to_datetime(e_val, errors='coerce')

                    if pd.notna(start_dt) and pd.notna(end_dt):
                        script_phases.append({
                            "Phase": phase_name,
                            "Start": start_dt.date(),
                            "End": end_dt.date(),
                            "Color": color
                        })
                except:
                    continue

        if script_phases:
            all_tasks.append({
                "UAT_Flag": uat_flag,
                "UAT_End_Date": uat_end_date,
                "Business Function Owner": str(row['Business Function Owner']),
                "Path": str(row['Path']),
                "Script Name": str(row['Script Name']),
                "Priority": row['Priority'],
                "Phases": script_phases
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt_With_UAT_End_Date"

    ws.merge_cells("A1:S1")
    title = ws.cell(1, 1, "GANTT CHART — UAT FLAGS WITH UAT END DATE")
    title.font = Font(name="Calibri", size=16, bold=True, color="FFFFFFFF")
    title.fill = PatternFill(start_color="FF4172B8", end_color="FF4172B8", fill_type="solid")
    title.alignment = Alignment(horizontal="center")

    headers = ["Business Function Owner", "Path", "Script Name", "Priority", "Phase", "UAT Flag",
               "UAT End Date", "Start", "End"]

    for i, h in enumerate(headers, 1):
        cell = ws.cell(3, i, h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF4172B8", end_color="FF4172B8", fill_type="solid")
        ws.column_dimensions[get_column_letter(i)].width = 20 if i <= 4 else 12

    all_starts = [phase["Start"] for task in all_tasks for phase in task["Phases"]]
    all_ends = [phase["End"] for task in all_tasks for phase in task["Phases"]]
    min_date = min(all_starts)
    max_date = max(all_ends)
    total_days = (max_date - min_date).days + 1

    timeline_start_col = len(headers) + 1
    for d in range(total_days):
        dt = min_date + timedelta(days=d)
        col = timeline_start_col + d
        cell = ws.cell(3, col, dt.strftime("%d\n%b"))
        cell.font = Font(bold=True, color="FFFFFFFF", size=9)

        if dt.weekday() >= 5:
            cell.fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FF4172B8", end_color="FF4172B8", fill_type="solid")

        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 4

    thin = Side(style="thin", color="FF000000")
    thick = Side(style="thick", color="FF000000")

    current_row = 4
    for task in all_tasks:
        phases = task["Phases"]
        n = len(phases)
        start_row = current_row
        end_row = current_row + n - 1

        merge_cols = {
            1: task["Business Function Owner"],
            2: task["Path"],
            3: task["Script Name"],
            4: task["Priority"]
        }

        for col_idx, value in merge_cols.items():
            if n > 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
            cell = ws.cell(start_row, col_idx, value)
            cell.alignment = Alignment(vertical="center", horizontal="left" if col_idx in [1, 2, 3] else "center", wrap_text=True)

        # Write phase rows including the new UAT End Date column
        for phase in phases:
            ws.cell(current_row, 5, phase["Phase"])
            ws.cell(current_row, 6, task["UAT_Flag"])
            ws.cell(current_row, 7, task["UAT_End_Date"])
            ws.cell(current_row, 8, phase["Start"])
            ws.cell(current_row, 9, phase["End"])

            for d in range(total_days):
                dt = min_date + timedelta(days=d)
                if phase["Start"] <= dt <= phase["End"]:
                    col = timeline_start_col + d
                    bar_cell = ws.cell(current_row, col, " ")

                    if dt.weekday() >= 5:
                        bar_cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
                    else:
                        bar_cell.fill = PatternFill(start_color=phase["Color"], end_color=phase["Color"], fill_type="solid")

            current_row += 1

        for r in range(start_row, end_row + 1):
            for c in range(1, timeline_start_col + total_days):
                cell = ws.cell(r, c)
                top = thick if r == start_row else thin
                bottom = thick if r == end_row else thin
                left = thick if c == 1 else thin
                right = thick if c == timeline_start_col + total_days - 1 else thin
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

        for c in range(1, timeline_start_col + total_days):
            sep_cell = ws.cell(end_row + 1, c)
            sep_cell.fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
        current_row += 1

    max_row = ws.max_row
    max_col = timeline_start_col + total_days - 1
    ws.freeze_panes = "H4"

    table_ref = f"A3:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="GanttTableWithUATEnd", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(OUTPUT_FILE)
    
    print(f"🎉 SUCCESS: Gantt chart with UAT End Date column created - {OUTPUT_FILE}")

if __name__ == "__main__":
    build_gantt_with_uat_end()
