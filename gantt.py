import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============== USER SETTINGS ==============
INPUT_FILE  = "Project_Scripts_Conversion_Tracking.xlsx"
SHEET_NAME  = "Sheet1"
OUTPUT_FILE = "Professional_Gantt_Fixed.xlsx"
# ===========================================

def build_gantt_from_excel_fixed():
    # Read your Excel file
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)

    # Required columns
    required = [
        "Business Function Owner", "Path", "Script Name", "Priority",
        "Analysis Start Date", "Analysis End Date",
        "Conversion Start Date", "Conversion End Date", 
        "Execution Start Date", "Execution End Date",
        "Recon Start Date", "Recon End Date",
        "Issue Fix Start Date", "Issue Fix End Date",
        "UAT Start Date", "UAT End Date"
    ]
    
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    # Add optional columns if missing
    for opt in ["Resource", "Percent Complete", "Status", "Notes"]:
        if opt not in df.columns:
            df[opt] = ""

    # FIXED: Proper color mapping with correct UAT color
    phase_colors = {
        "Analysis": "FF5B9BD5",     # Blue
        "Conversion": "FF70AD47",   # Green
        "Execution": "FFED7D31",    # Orange
        "Recon": "FFA5A5A5",        # Gray
        "Issue Fix": "FFFFC000",    # Yellow
        "UAT": "FF4472C4"           # Dark Blue - FIXED
    }

    phases_map = [
        ("Analysis", "Analysis Start Date", "Analysis End Date"),
        ("Conversion", "Conversion Start Date", "Conversion End Date"),
        ("Execution", "Execution Start Date", "Execution End Date"),
        ("Recon", "Recon Start Date", "Recon End Date"),
        ("Issue Fix", "Issue Fix Start Date", "Issue Fix End Date"),
        ("UAT", "UAT Start Date", "UAT End Date")
    ]

    # FIXED: Proper grouping by all three identifiers
    grouped_items = []
    for keys, group in df.groupby(['Business Function Owner', 'Path', 'Script Name'], sort=False):
        bfo, path, script = keys
        row0 = group.iloc[0]
        
        item = {
            "Business Function Owner": str(bfo),  # FIXED: Ensure string conversion
            "Path": str(path),
            "Script Name": str(script),
            "Priority": row0.get("Priority", ""),
            "Resource": row0.get("Resource", ""),
            "Percent Complete": row0.get("Percent Complete", ""),
            "Status": row0.get("Status", ""),
            "Notes": row0.get("Notes", ""),
            "Phases": []
        }

        # Extract all phases for this group
        for phase_name, s_col, e_col in phases_map:
            s_val = row0.get(s_col)
            e_val = row0.get(e_col)
            if pd.notna(s_val) and pd.notna(e_val) and s_val != "########" and e_val != "########":
                try:
                    start_date = pd.to_datetime(s_val).date()
                    end_date = pd.to_datetime(e_val).date()
                    if end_date >= start_date:
                        item["Phases"].append({
                            "Phase": phase_name,
                            "Start": start_date,
                            "End": end_date,
                            "Color": phase_colors[phase_name]  # FIXED: Ensures UAT gets correct color
                        })
                except Exception as e:
                    print(f"Date parsing error for {phase_name}: {e}")
                    continue

        if item["Phases"]:
            grouped_items.append(item)

    if not grouped_items:
        raise ValueError("No valid phases found.")

    # Timeline calculation
    all_starts = [p["Start"] for item in grouped_items for p in item["Phases"]]
    all_ends = [p["End"] for item in grouped_items for p in item["Phases"]]
    min_date = min(all_starts)
    max_date = max(all_ends)
    total_days = (max_date - min_date).days + 1

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Professional_Gantt"

    # Title
    ws.merge_cells("A1:R1")
    title = ws.cell(1, 1, "PROJECT GANTT CHART — BUSINESS FUNCTION OWNER / PATH / SCRIPT")
    title.font = Font(name="Calibri", size=18, bold=True, color="FFFFFFFF")
    title.fill = PatternFill(start_color="FF2F5597", end_color="FF2F5597", fill_type="solid")
    title.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Business Function Owner", "Path", "Script Name", "Priority", "Resource",
        "Phase", "Start", "End", "% Complete", "Status", "Notes"
    ]
    
    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        # Set wider columns for text fields
        width = 28 if h in ["Business Function Owner", "Path", "Script Name", "Notes"] else 14
        ws.column_dimensions[get_column_letter(i)].width = width

    # Timeline headers
    timeline_start_col = len(headers) + 1
    for d in range(total_days):
        dt = min_date + timedelta(days=d)
        col = timeline_start_col + d
        cell = ws.cell(3, col, dt.strftime("%d\n%b"))
        cell.font = Font(bold=True, color="FFFFFFFF", size=9)
        cell.fill = PatternFill(start_color="FF5B9BD5", end_color="FF5B9BD5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 4

    # Border style
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write data rows
    current_row = 4
    for item in grouped_items:
        phases = item["Phases"]
        n = len(phases)
        start_block = current_row
        end_block = current_row + n - 1

        # FIXED: Merge the three main identifiers
        merge_columns = {
            1: item["Business Function Owner"],  # FIXED: Now properly writes BFO
            2: item["Path"],
            3: item["Script Name"],
            4: item["Priority"],
            5: item["Resource"]
        }

        for col_idx, value in merge_columns.items():
            if n > 1:  # Only merge if multiple phases
                ws.merge_cells(start_row=start_block, start_column=col_idx, 
                             end_row=end_block, end_column=col_idx)
            
            cell = ws.cell(start_block, col_idx, value)
            align_h = "left" if col_idx in (1, 2, 3) else "center"
            cell.alignment = Alignment(vertical="center", horizontal=align_h, wrap_text=True)

        # Write phase rows
        for phase in phases:
            ws.cell(current_row, 6, phase["Phase"])
            ws.cell(current_row, 7, phase["Start"])
            ws.cell(current_row, 8, phase["End"])
            ws.cell(current_row, 9, item["Percent Complete"])
            
            # Status with color coding
            status_cell = ws.cell(current_row, 10, item["Status"])
            status_colors = {
                "Completed": "FF70AD47",
                "In Progress": "FFED7D31", 
                "Delayed": "FFFF0000"
            }
            status_fill = status_colors.get(str(item["Status"]), "FFD9D9D9")
            status_cell.fill = PatternFill(start_color=status_fill, end_color=status_fill, fill_type="solid")
            if status_fill in ("FF70AD47", "FFED7D31", "FFFF0000"):
                status_cell.font = Font(bold=True, color="FFFFFFFF")
            
            ws.cell(current_row, 11, item["Notes"])

            # FIXED: Gantt bars with correct colors including UAT
            for d in range(total_days):
                dt = min_date + timedelta(days=d)
                if phase["Start"] <= dt <= phase["End"]:
                    col = timeline_start_col + d
                    bar_cell = ws.cell(current_row, col, " ")
                    bar_cell.fill = PatternFill(start_color=phase["Color"], 
                                               end_color=phase["Color"], fill_type="solid")

            current_row += 1

        # Add spacer row between script blocks
        for c in range(1, timeline_start_col + total_days):
            ws.cell(current_row, c).fill = PatternFill(start_color="FFF8F8F8", 
                                                      end_color="FFF8F8F8", fill_type="solid")
        current_row += 1

    # Apply borders
    max_row = ws.max_row
    max_col = timeline_start_col + total_days - 1
    for r in range(3, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).border = border

    # Weekend shading
    for d in range(total_days):
        dt = min_date + timedelta(days=d)
        if dt.weekday() >= 5:  # Weekend
            col = timeline_start_col + d
            for r in range(4, max_row + 1):
                cell = ws.cell(r, col)
                # Only shade if no Gantt bar already there
                if not cell.fill or cell.fill.start_color.index == "00000000":
                    cell.fill = PatternFill(start_color="FFF2F2F2", 
                                          end_color="FFF2F2F2", fill_type="solid")

    # Today marker
    today = datetime.now().date()
    if min_date <= today <= max_date:
        today_col = timeline_start_col + (today - min_date).days
        for r in range(3, max_row + 1):
            ws.cell(r, today_col).border = Border(left=Side(style="thick", color="FFFF0000"))

    # Excel Table for filtering
    table_ref = f"A3:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="GanttTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, 
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Legend
    legend_row = max_row + 2
    ws.cell(legend_row, 1, "LEGEND").font = Font(size=14, bold=True)
    legend_items = [
        ("FF5B9BD5", "Analysis"),
        ("FF70AD47", "Conversion"), 
        ("FFED7D31", "Execution"),
        ("FFA5A5A5", "Recon"),
        ("FFFFC000", "Issue Fix"),
        ("FF4472C4", "UAT"),  # FIXED: UAT legend
        ("FFFF0000", "Today (red line)"),
        ("FFF2F2F2", "Weekends")
    ]
    
    for i, (color, label) in enumerate(legend_items, start=1):
        cell = ws.cell(legend_row + i, 1, label)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF")
        ws.column_dimensions["A"].width = 40

    # Freeze panes
    ws.freeze_panes = "F4"  # Keeps identifiers visible while scrolling

    wb.save(OUTPUT_FILE)
    print(f"✅ Professional Gantt Chart created: {OUTPUT_FILE}")
    return OUTPUT_FILE

if __name__ == "__main__":
    build_gantt_from_excel_fixed()
